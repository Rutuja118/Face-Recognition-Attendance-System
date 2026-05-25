print("Program started")

import tkinter as tk
from tkinter import messagebox
import threading
import os
import cv2
import face_recognition
import numpy as np
from datetime import datetime
import pandas as pd
import time
import mysql.connector
from openpyxl import load_workbook
from openpyxl.styles import Font

# ================== MYSQL CONNECTION ==================
conn = mysql.connector.connect(
    host="localhost",
    user="root",
    password="password",
    database="face_attendance"
)

if conn.is_connected():
    print("✅ MySQL Connected Successfully")

cursor = conn.cursor()

# ================== LOGIN SYSTEM ==================
USERNAME = "user"
PASSWORD = "1234"

def login():

    user = username_entry.get()
    pwd = password_entry.get()

    if user == USERNAME and pwd == PASSWORD:

        login_window.destroy()

        open_main_app()

    else:

        messagebox.showerror(
            "Login Failed",
            "Invalid Username or Password"
        )

# ================== MAIN VARIABLES ==================
running = False
cap = None
last_unknown_time = 0

attendance_count = 0

last_detected_name = ""
last_detection_time = 0

# ================== TOTAL ATTENDANCE ==================
def update_total_attendance():

    global attendance_count

    try:

        cursor.execute("SELECT COUNT(*) FROM attendance")

        attendance_count = cursor.fetchone()[0]

        count_label.config(
            text=f"Total Attendance: {attendance_count}"
        )

    except Exception as e:

        print("Error:", e)

# ================== ATTENDANCE FUNCTION ==================
def markAttendance(name):

    global status_label
    global attendance_count

    now = datetime.now()

    # ✅ Proper Date and Time
    dateString = now.strftime('%Y-%m-%d')
    timeString = now.strftime('%H:%M:%S')

    # ✅ Check if already marked today
    query = """
    SELECT * FROM attendance
    WHERE name=%s AND date=%s
    """

    cursor.execute(query, (name, dateString))

    result = cursor.fetchone()

    # ✅ Already marked
    if result:

        root.after(
            0,
            lambda: status_label.config(
                text=f"⚠️ {name} already marked today",
                fg="orange"
            )
        )

        return

    # ✅ Insert attendance
    insert_query = """
    INSERT INTO attendance(name, date, time)
    VALUES(%s, %s, %s)
    """

    cursor.execute(
        insert_query,
        (name, dateString, timeString)
    )

    conn.commit()

    attendance_count += 1

    print(f"✅ Attendance Marked for {name}")

    root.after(
        0,
        lambda: status_label.config(
            text=f"✅ Attendance Marked for {name}",
            fg="green"
        )
    )

    root.after(
        0,
        lambda: count_label.config(
            text=f"Total Attendance: {attendance_count}"
        )
    )

# ================== EXPORT TO EXCEL ==================
def convert_to_excel():

    try:

        query = "SELECT * FROM attendance"

        df = pd.read_sql(query, conn)

        # ✅ Convert properly
        df['date'] = df['date'].astype(str)
        df['time'] = df['time'].apply(lambda x: str(x).split()[-1])

        file_name = "attendance.xlsx"

        # ✅ Save Excel
        with pd.ExcelWriter(
            file_name,
            engine='openpyxl'
        ) as writer:

            df.to_excel(
                writer,
                index=False
            )

            worksheet = writer.sheets['Sheet1']

            # ✅ Column Width
            worksheet.column_dimensions['A'].width = 10
            worksheet.column_dimensions['B'].width = 25
            worksheet.column_dimensions['C'].width = 20
            worksheet.column_dimensions['D'].width = 20

        # ✅ Bold Header
        wb = load_workbook(file_name)

        ws = wb.active

        for cell in ws[1]:

            cell.font = Font(bold=True)

        wb.save(file_name)

        print("✅ Excel Exported Successfully")

    except Exception as e:

        print("❌ Excel Error:", e)

# ================== FACE ENCODING ==================
path = 'images'

images = []
classNames = []

for cl in os.listdir(path):

    img = cv2.imread(f'{path}/{cl}')

    if img is not None:

        images.append(img)

        name= os.path.splitext(cl)[0]
          
          #Remove image number
        name=name.split("_")[0]
        classNames.append(name)
          
        

def findEncodings(images):

    encodeList = []

    for img in images:

        img = cv2.cvtColor(
            img,
            cv2.COLOR_BGR2RGB
        )

        encodes = face_recognition.face_encodings(img)

        if len(encodes) > 0:

            encodeList.append(encodes[0])

    return encodeList

encodeListKnown = findEncodings(images)

print("✅ Encoding Complete")

# ================== CAMERA ==================
def start_camera():

    global running
    global cap
    global last_unknown_time
    global last_detected_name
    global last_detection_time

    running = True

    cap = cv2.VideoCapture(0)

    while running:

        success, img = cap.read()

        if not success:

            break

        imgS = cv2.resize(
            img,
            (0, 0),
            None,
            0.25,
            0.25
        )

        imgS = cv2.cvtColor(
            imgS,
            cv2.COLOR_BGR2RGB
        )

        facesCurFrame = face_recognition.face_locations(imgS)

        encodesCurFrame = face_recognition.face_encodings(
            imgS,
            facesCurFrame
        )

        for encodeFace, faceLoc in zip(
            encodesCurFrame,
            facesCurFrame
        ):

            faceDis = face_recognition.face_distance(
                encodeListKnown,
                encodeFace
            )

            matchIndex = np.argmin(faceDis)

            # ================== KNOWN FACE ==================
            if faceDis[matchIndex] < 0.6:

                name = classNames[matchIndex].upper()

                current_time = time.time()

                # ✅ Prevent duplicate scanning
                if (
                    name != last_detected_name
                    or current_time - last_detection_time > 3
                ):

                    last_detected_name = name
                    last_detection_time = current_time

                    markAttendance(name)

            # ================== UNKNOWN FACE ==================
            else:

                name = "UNKNOWN"

                current_time = time.time()

                if current_time - last_unknown_time > 3:

                    last_unknown_time = current_time

                    root.after(
                        0,
                        lambda: status_label.config(
                            text="❌ Unknown Face Detected",
                            fg="red"
                        )
                    )

            # ================== FACE BOX ==================
            y1, x2, y2, x1 = faceLoc

            y1 *= 4
            x2 *= 4
            y2 *= 4
            x1 *= 4

            color = (0, 255, 0)

            if name == "UNKNOWN":

                color = (0, 0, 255)

            cv2.rectangle(
                img,
                (x1, y1),
                (x2, y2),
                color,
                2
            )

            cv2.putText(
                img,
                name,
                (x1, y1 - 10),
                cv2.FONT_HERSHEY_SIMPLEX,
                1,
                (255, 255, 255),
                2
            )

        cv2.imshow("Camera", img)

        if cv2.waitKey(1) & 0xFF == ord('q'):

            stop_camera()

            break

    cap.release()

    cv2.destroyAllWindows()

    # ✅ Export Excel
    convert_to_excel()

# ================== STOP CAMERA ==================
def stop_camera():

    global running

    running = False

# ================== OPEN EXCEL ==================
def open_excel():

    try:

        # ✅ Update latest data before opening
        convert_to_excel()

        os.startfile("attendance.xlsx")

    except:

        messagebox.showerror(
            "Error",
            "Excel file not found!"
        )

# ================== GUI ==================
def open_main_app():

    global root
    global status_label
    global count_label

    root = tk.Tk()

    root.title("Face Attendance System")

    root.geometry("400x400")

    root.config(bg="lightblue")

    tk.Label(
        root,
        text="Face Attendance System",
        font=("Arial", 16, "bold"),
        bg="lightblue"
    ).pack(pady=20)

    # ✅ Status Label
    status_label = tk.Label(
        root,
        text="Waiting for face scan...",
        font=("Arial", 12),
        bg="lightblue"
    )

    status_label.pack(pady=10)

    # ✅ Total Attendance Label
    count_label = tk.Label(
        root,
        text="Total Attendance: 0",
        font=("Arial", 12, "bold"),
        bg="lightblue"
    )

    count_label.pack(pady=10)

    tk.Button(
        root,
        text="Start Camera",
        command=lambda: threading.Thread(
            target=start_camera
        ).start(),
        width=20,
        bg="green",
        fg="white"
    ).pack(pady=10)

    tk.Button(
        root,
        text="Stop Camera",
        command=stop_camera,
        width=20,
        bg="red",
        fg="white"
    ).pack(pady=10)

    tk.Button(
        root,
        text="Open Excel",
        command=open_excel,
        width=20,
        bg="blue",
        fg="white"
    ).pack(pady=10)

    tk.Button(
        root,
        text="Exit",
        command=root.destroy,
        width=20,
        bg="black",
        fg="white"
    ).pack(pady=10)

    # ✅ Load total attendance count
    update_total_attendance()

    root.mainloop()

# ================== LOGIN WINDOW ==================
login_window = tk.Tk()

login_window.title("Login")

login_window.geometry("300x250")

tk.Label(
    login_window,
    text="Login System",
    font=("Arial", 14, "bold")
).pack(pady=10)

tk.Label(
    login_window,
    text="Username"
).pack()

username_entry = tk.Entry(login_window)

username_entry.pack()

tk.Label(
    login_window,
    text="Password"
).pack()

password_entry = tk.Entry(
    login_window,
    show="*"
)

password_entry.pack()

tk.Button(
    login_window,
    text="Login",
    command=login,
    bg="blue",
    fg="white"
).pack(pady=20)

login_window.mainloop()
